# 다음 영화 순위 크롤링 하는 프로그램
import requests
from bs4 import BeautifulSoup
import csv

# 다음 영화 순위 페이지 URL
url = "https://movie.daum.net/ranking/reservation"

# HTTP 요청 보내기
headers = {'User-Agent' : 'Mozilla/5.0 (Windows NT 10.0; Win64; x64)AppleWebKit/537.36 (KHTML, like Gecko) Chrome/73.0.3683.86 Safari/537.36'}
response = requests.get(url, headers=headers)

# HTTP 요청이 성공했는지 확인하기
if response.status_code == 200:
    # HTML 파싱하기
    soup = BeautifulSoup(response.text, "html.parser")  

# 엑셀에 데이터 쓰기
from openpyxl import Workbook

# 엑셀파일 쓰기
write_wb = Workbook()

# 이름이 있는 시트를 생성
write_ws = write_wb.create_sheet('다음 영화 순위')  

# Sheet1에다 입력 
write_ws = write_wb.active
write_ws['A1'] = '순위'
write_ws['B1'] = '제목'
write_ws['C1'] = '평점'
write_ws['D1'] = '예매율'
write_ws['E1'] = '개봉날짜'

#행 단위로 추가 순위, 제목, 평점, 예매율, 개봉날짜
rank = 0
movie_list = soup.select(".thumb_cont")
for tr in movie_list:
    rank = rank + 1
    a_tag = tr.select_one("a")
    txt_grade = tr.select_one("span.txt_grade")
    txt_num = tr.select_one("span.txt_num")
    txt_date = tr.select_one(".txt_info > span.txt_num")
    write_ws.append([f'{rank}',(f'{a_tag.text}'),f'{txt_grade.text}',f'{txt_num.text}',f'{txt_date.text}'])

#셀 단위로 추가
#write_ws.cell(5, 5, '5행5열')

write_wb.save('다음 영화 순위.xlsx')